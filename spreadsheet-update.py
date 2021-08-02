import openpyxl
from openpyxl import cell

orderfile=openpyxl.load_workbook("orders.xlsx")


customer_loc=orderfile["Sheet1"]
customer_phone=orderfile["Sheet2"]
customer_orders=orderfile["Sheet3"]

for customer in range(2,customer_loc.max_row+1):
    print(customer_loc.cell(customer,2).value)
    id=int(customer_loc.cell(customer,1).value) 
    
    ##Update phone number on Sheet1
    
    for phone in range(2,customer_phone.max_row+1):
        if ( id == int(customer_phone.cell(phone,1).value)) :
            customer_loc.cell(customer,5).value=int(customer_phone.cell(phone,2).value)
    
    ##Update total purchase on Sheet1 and total cost on Sheet3
    total_purchase = 0            
    for order in range(2,customer_orders.max_row+1):
        total_cost= int( customer_orders.cell(order,3).value * customer_orders.cell(order,4).value)
        if ( id == int(customer_orders.cell(order,2).value)) :
            total_purchase = total_purchase + total_cost
        
        customer_orders.cell(order,5).value = total_cost
    customer_loc.cell(customer,6).value = total_purchase
    
        

orderfile.save("solution.xlsx")